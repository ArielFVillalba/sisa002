from django.shortcuts import render, HttpResponse, redirect , get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils.datetime_safe import datetime
from django.views.generic import CreateView, UpdateView, DeleteView
from openpyxl.compat import numbers
from openpyxl.styles import Alignment

from .forms import *
from django.contrib import messages
from reportlab.lib import colors
from django.http import JsonResponse, FileResponse, HttpResponse
from django.shortcuts import render
from inicio.funcion import *
from django.shortcuts import redirect
import io
import csv
from reportlab.pdfgen import canvas
import time
import datetime
from datetime import datetime

from reportlab.lib.pagesizes import letter, landscape, inch, A4, A3, legal, B4, B5, A5, portrait
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse


def presupdeventacabinf_listar(request):
    if not request.user.is_authenticated:
        return redirect('login')

    idcliente = request.POST['idcliente']
    fechaini = request.POST['fechaini']
    fechafin = request.POST['fechafin']
    tipodoc = request.POST['tipodoc']
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab = Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if idcliente == '':
        idcliente = 0
    if presupdeventacab.exists() and int(idcliente) >= 1:
        presupdeventacab = presupdeventacab .filter(idcliente_id=idcliente)
    if presupdeventacab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab= presupdeventacab .filter(fecha__gte=fecha_inicio)
    if presupdeventacab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab= presupdeventacab .filter(fecha__lte=fecha_fin)
    if presupdeventacab.exists() and tipodoc:
        presupdeventacab= presupdeventacab .filter(tipodoc=tipodoc)

    datos = []
    for objeto in presupdeventacab :
        datos.append({
        'fecha': objeto.fecha.strftime("%d/%m/%Y"),
        'nropresupuesto': objeto.nropresupuesto,
        'cliente': objeto.cliente,
        'tipodoc': objeto.tipodoc,
        'total':objeto.total
    })
    total = sum([objeto.total * objeto.total for objeto in presupdeventacab ])
    Response = JsonResponse({'datos': datos})
    return Response



def presupdeventacab_inf(request):
    if not request.user.is_authenticated:
        return redirect('login')

    return render(request, 'ventasinf/lstpresupdeventa.html')

def presupdeventacabinfcsv(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # -----------------------------------------------
    idcliente = request.GET.get('idcliente', '')
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    tipodoc = request.GET.get('tipodoc', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab = Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)

    if idcliente == '':
        idcliente = 0
    if presupdeventacab.exists() and int(idcliente) >= 1:
        presupdeventacab= presupdeventacab .filter(idcliente_id=idcliente)
    if presupdeventacab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab= presupdeventacab .filter(fecha__gte=fecha_inicio)
    if presupdeventacab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab= presupdeventacab .filter(fecha__lte=fecha_fin)
    if presupdeventacab.exists() and tipodoc:
        presupdeventacab= presupdeventacab .filter(tipodoc=tipodoc)

    # -------------------------------------------------

    objetos=presupdeventacab 

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="datos.csv"'
    writer = csv.writer(response, delimiter=';')  # Establece el separador como ";"
    writer.writerow(["", "", "PRESUPUESTO ", "",""])

    writer.writerow(["FECHA", "NRO PRESUPUESTO", "CLIENTE", "TIPODOC","TOTAL"])
    # Agrega los datos de cada fila
    for fila in objetos:
        writer.writerow([fila.fecha, fila.nropresupuesto, fila.cliente, fila.tipodoc, fila.total])
    return response

def presupdeventadetinfcsv(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # -----------------------------------------------
    idcliente = request.GET.get('idcliente', '')
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    tipodoc = request.GET.get('tipodoc', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab = Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if idcliente == '':
        idcliente = 0
    if presupdeventacab.exists() and int(idcliente) >= 1:
        presupdeventacab= presupdeventacab.filter(idcliente_id=idcliente)
    if presupdeventacab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab= presupdeventacab.filter(fecha__gte=fecha_inicio)
    if presupdeventacab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab= presupdeventacab.filter(fecha__lte=fecha_fin)
    if presupdeventacab.exists() and tipodoc:
        presupdeventacab= presupdeventacab.filter(tipodoc=tipodoc)

    # -------------------------------------------------

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="datos.csv"'
    writer = csv.writer(response, delimiter=';')  # Establece el separador como ";"
    writer.writerow(["", "", "INFORME DE PRESUPUESTO DE VENTA DETALLADO ","","","" ,"","","",""])

    writer.writerow(["FECHA", "FACTURA", "CLIENTE","ORDEN","codigo","descripcion" ,"cantidad","precio","iva","subtotal"])
    total=0
    for objetocab in presupdeventacab :
        didpresupuestoventacab = objetocab.idpresupuestoventacab
        presupdeventadet = Presupuestoventadet.objects.filter(idpresupuestoventacab =didpresupuestoventacab).order_by('orden')

        for objetodet in presupdeventadet:
            if objetodet.orden==1 and total>0:
                    writer.writerow(["","","","","","","","","TOTAL",formatnumver(total,0)])
                    total = 0
                    writer.writerow([""])
            total=total+(objetodet.cantidad * objetodet.precio)

            writer.writerow([
                objetocab.fecha.strftime("%d/%m/%Y"),
                objetocab.nropresupuesto,
                objetocab.cliente,
                objetodet.orden,
                objetodet.codigo,
                objetodet.descripcion,
                formatnumver(objetodet.cantidad,3),
                formatnumver(objetodet.precio,0),
                formatnumver(objetodet.iva,0),
                formatnumver((objetodet.cantidad * objetodet.precio), 0)
            ])
    writer.writerow(["", "", "", "", "", "", "", "", "TOTAL", formatnumver(total, 0)])

    return response



def presupdeventacabinfexcel(request):
    # -----------------------------------------------
    idcliente = request.GET.get('idcliente', '')
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    tipodoc = request.GET.get('tipodoc', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab = Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if idcliente == '':
        idcliente = 0
    if presupdeventacab.exists() and int(idcliente) >= 1:
        presupdeventacab= presupdeventacab .filter(idcliente_id=idcliente)
    if presupdeventacab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab= presupdeventacab .filter(fecha__gte=fecha_inicio)
    if presupdeventacab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab = presupdeventacab .filter(fecha__lte=fecha_fin)
    if presupdeventacab.exists() and tipodoc:
        presupdeventacab   = presupdeventacab .filter(tipodoc=tipodoc)

    # -------------------------------------------------
    data = presupdeventacab 

    # Crear un nuevo libro de trabajo (workbook)
    wb = Workbook()

    # Obtén tus datos de Django y genera las filas del archivo Excel
    # Ejemplo ficticio con datos de un modelo llamado "MiModelo"

    # Crea una hoja de cálculo en el libro de trabajo
    sheet = wb.active
    sheet.title = "'NOTA DEBDITO DE VENTA '"

    # Agrega encabezados a la hoja de cálculo
    # headers = ['Campo1', 'Campo2', 'Campo3']
    headers=["FECHA", "NRO PRESUPUESTO", "CLIENTE", "TIPODOC","TOTAL"]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}2"]
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alinea el título al centro
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15


    # Fusiona las celdas para el título general
    inicio_celda = sheet.cell(row=1, column=1)
    fin_celda = sheet.cell(row=1, column=len(headers))
    sheet.merge_cells(start_row=1, start_column=inicio_celda.column, end_row=1, end_column=fin_celda.column)
    inicio_celda.alignment = Alignment(horizontal='center')
    # Establece el valor del título general en la celda fusionada
    titulo_general = "LISTADO PRESUPUESTO VENTA"
    inicio_celda.value = titulo_general

    # Agrega los datos de cada fila a la hoja de cálculo
    for row_num, fila in enumerate(data, 3):
        sheet.cell(row=row_num, column=1, value=fila.fecha.strftime('%d/%m/%Y'))
        sheet.cell(row=row_num, column=2, value=fila.nropresupuesto)
        sheet.cell(row=row_num, column=3, value=fila.cliente)
        sheet.cell(row=row_num, column=4, value=fila.tipodoc)
        total_cell = sheet.cell(row=row_num, column=5, value=fila.total)


    # Configura el tipo de respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="datos.xlsx"'

    # Guarda el libro de trabajo en la respuesta HTTP
    wb.save(response)

    return response

def presupdeventadetinfexcel(request):
    # -----------------------------------------------
    idcliente = request.GET.get('idcliente', '')
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    tipodoc = request.GET.get('tipodoc', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab   = Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if idcliente == '':
        idcliente = 0
    if presupdeventacab .exists() and int(idcliente) >= 1:
        presupdeventacab   = presupdeventacab .filter(idcliente_id=idcliente)
    if presupdeventacab .exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab   = presupdeventacab .filter(fecha__gte=fecha_inicio)
    if presupdeventacab .exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab   = presupdeventacab .filter(fecha__lte=fecha_fin)
    if presupdeventacab .exists() and tipodoc:
        presupdeventacab   = presupdeventacab .filter(tipodoc=tipodoc)

    # -------------------------------------------------

    # Crear un nuevo libro de trabajo (workbook)
    wb = Workbook()

    # Obtén tus datos de Django y genera las filas del archivo Excel
    # Ejemplo ficticio con datos de un modelo llamado "MiModelo"

    # Crea una hoja de cálculo en el libro de trabajo
    sheet = wb.active
    sheet.title = "LISTADO_DETALLADO"

    # Agrega encabezados a la hoja de cálculo
    # headers = ['Campo1', 'Campo2', 'Campo3']
    headers=["FECHA", "FACTURA", "CLIENTE", "ORDEN","CODIGO", "DESCRIPCION", "CANTIDAD","PRECIO", "IVA","SUBTOTAL"]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}2"]
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alinea el título al centro
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15


    # Fusiona las celdas para el título general
    inicio_celda = sheet.cell(row=1, column=1)
    fin_celda = sheet.cell(row=1, column=len(headers))
    sheet.merge_cells(start_row=1, start_column=inicio_celda.column, end_row=1, end_column=fin_celda.column)
    inicio_celda.alignment = Alignment(horizontal='center')
    # Establece el valor del título general en la celda fusionada
    titulo_general = "LISTADO PRESUPUESTO DE  VENTAS DETALLADO  "
    inicio_celda.value = titulo_general
    row_num=3
    total=0
    for objetocab in presupdeventacab :
        didpresupuestoventacab = objetocab.idpresupuestoventacab
        presupdeventadet = Presupuestoventadet.objects.filter(idpresupuestoventacab =didpresupuestoventacab).order_by('orden')
        if total>0:
            row_num=row_num+1
            sheet.cell(row=row_num, column=10, value=formatnumver(total,0))
            total=0
            row_num=row_num+1

        for objetodet in presupdeventadet:
            total=total+(objetodet.cantidad * objetodet.precio)
            row_num=row_num+1
            sheet.cell(row=row_num, column=1, value=objetocab.fecha.strftime('%d/%m/%Y'))
            sheet.cell(row=row_num, column=2, value=objetocab.nropresupuesto)
            sheet.cell(row=row_num, column=3, value=objetocab.cliente)
            sheet.cell(row=row_num, column=4, value=objetodet.orden)
            sheet.cell(row=row_num, column=5, value=objetodet.codigo)
            sheet.cell(row=row_num, column=6, value=objetodet.descripcion)
            sheet.cell(row=row_num, column=7, value=formatnumver(objetodet.cantidad,3))
            sheet.cell(row=row_num, column=8, value=formatnumver(objetodet.precio,0))
            sheet.cell(row=row_num, column=9, value=formatnumver(objetodet.iva,0))
            sheet.cell(row=row_num, column=10, value=formatnumver((objetodet.cantidad * objetodet.precio),0))
    row_num = row_num + 1
    sheet.cell(row=row_num, column=10, value=formatnumver(total, 0))

    # Configura el tipo de respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="datos.xlsx"'

    # Guarda el libro de trabajo en la respuesta HTTP
    wb.save(response)

    return response

def presupdeventacabinfpdf(request):
    # -----------------------------------------------
    idcliente = request.GET.get('idcliente', '')
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    tipodoc = request.GET.get('tipodoc', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab= Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if idcliente == '':
        idcliente = 0
    if presupdeventacab.exists() and int(idcliente) >= 1:
        presupdeventacab = presupdeventacab .filter(idcliente_id=idcliente)
    if presupdeventacab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab = presupdeventacab .filter(fecha__gte=fecha_inicio)
    if presupdeventacab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab = presupdeventacab .filter(fecha__lte=fecha_fin)
    if presupdeventacab.exists() and tipodoc:
        presupdeventacab= presupdeventacab .filter(tipodoc=tipodoc)

    # -------------------------------------------------

    #letter=8.5 * 11 pulgadas  215.9 *279.4 mm
    buf = io.BytesIO()
    #c = canvas.Canvas(buf)
    custom_oficio_width = 216 *2.7 # Anchura en unidades de medida (por ejemplo, puntos)
    custom_oficio_height = 330*2.1 # Altura en unidades de medida (por ejemplo, puntos)
    pagesize_oficio = portrait((custom_oficio_width, custom_oficio_height))
    c = canvas.Canvas(buf, pagesize=pagesize_oficio)
    #c = canvas.Canvas(buf, pagesize=legal)

    presupdeventacab_infpdftit(c,680,0)
    presupdeventacab_infpdfdet(presupdeventacab ,c,620)
    c.save()
    buf.seek(0)

    return FileResponse(buf, filename='venuepdf.pdf')

gancho_col = [80, 80, 100, 60, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
galto_fila = 20  # Alto de cada fila de la tabla
gtipo_col = ["t", "t", "t","t","n"]  # tipo de cada columna de la tabla
galin_col = ["l", "l", "l","l","r"]  # tipo de cada columna de la tabla
gcantcar_col = [10, 12, 20,20,16]  # tipo de cada columna de la tabla
gdec_col = [0, 0, 0,0,3]  # tipo de cada columna de la tabla

def presupdeventacab_infpdftit(c,y,npag):
        ahora = time.strftime("%c")
        fechahora = time.strftime("%c")
        fecha = time.strftime("%x")
        print(npag)

        #x = datetime.datetime.now()
        lx = 100
        hi = y
        c.setFont('Helvetica', 12)
        #c.drawString(lx+300, hi , "usuario:  fecha : " + fecha + "   " + str(x.hour) + ":" + str(
        #    x.minute) + "   pagina : " + str(nropag))
        c.setFont('Helvetica', 16)

         # Definir el color de relleno y el color de letra deseado
        color_relleno = (0.9, 0.9, 0.9)  # Utiliza los valores RGB del color de relleno que desees
        color_letra = (1, 0, 0)  # Utiliza los valores RGB del color de letra rojo
        color_relleno=(0, 0, 1)  # Azul en formato RGB
        color_letra=(1, 1, 1)  # blanco en formato RGB
        color_celeste=(0.529, 0.808, 0.922)  # celeste en formato RGB


        # Dibujar la cuadrícula
        ancho =150
        alto=30
        xlist = [lx, lx + ancho]
        ylist = [hi , hi + alto]
        # c.grid(xlist, ylist)

        # Establecer el color de relleno y dibujar el rectángulo en una celda específica
        c.setFillColor(colors.white)
        c.rect(lx+1, hi+1,  ancho-2, alto-2, fill=True, stroke=False)

        # Establecer el color de letra y dibujar el título en la celda
        #c.setFillColorRGB(*color_letra)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 22)  # Puedes ajustar la fuente y el tamaño según tus necesidades
        c.drawString(lx + 20, hi-20,"INFORME PRESUPUESTO DE VENTA ")
        presupdeventacab_infpdfcabdet(c,620)


def presupdeventacab_infpdfcabdet(c,y):
    data = [["FECHA", "N ORDEN", "CLIENTE", "TIPODOC","TOTAL"]]

    # Definir coprespadas y dimensiones de la tabla
    fila_x = 20  # Coprespada x de la esquina superior izquierda de la tabla
    fila_y = y  # Coprespada y de la esquina superior izquierda de la tabla
    alto_fila = 20  # Alto de cada fila de la tabla
    ancho_col = [80, 80, 100, 60, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
    tipo_col = ["t", "t", "t", "t", "t"]  # tipo de cada columna de la tabla
    alin_col = ["l", "l", "l", "l", "r"]  # tipo de cada columna de la tabla
    cantcar_col = [10, 12, 20, 20, 16]  # tipo de cada columna de la tabla
    color_celeste = (0.529, 0.808, 0.922)  # celeste en formato RGB
    # Dibujar la tabla
    for i, fila in enumerate(data):
        for j, valor in enumerate(fila):
            xlist = [fila_x - 1, fila_x + ancho_col[j]]
            ylist = [fila_y, fila_y + alto_fila]
            c.grid(xlist, ylist)
            c.setFillColor(color_celeste)  # Establecer el color de fondo de la celda
            c.setFont("Helvetica", 12)  # Puedes ajustar la fuente y el tamaño según tus necesidades

            valorw = formatearcampo(str(valor), int(cantcar_col[j]), alin_col[j], tipo_col[j])
            c.drawString(fila_x + 3, fila_y + alto_fila / 4, str(valorw))  # Agregar el valor a la celda
            fila_x = fila_x + ancho_col[j]


def presupdeventacab_infpdfdet(presupdeventacab ,c,y):
        total=0
        detmax=25
        nropag=1
        #data = [["FECHA", "FACTURA", "CLIENTE", "TIPODOC","TOTAL"]]
        data = []
        objetos = presupdeventacab 
        for objeto in objetos:
                nueva_fila = [str(objeto.fecha.strftime('%d/%m/%Y')),
                str(objeto.nropresupuesto),
                str(objeto.cliente),
                str(objeto.tipodoc),
                objeto.total
                ]
                data.append(nueva_fila)

        # Definir coprespadas y dimensiones de la tabla
        fila_x = 30  # Coprespada x de la esquina superior izquierda de la tabla
        fila_y = y  # Coprespada y de la esquina superior izquierda de la tabla
        ancho_col = [80, 80, 100, 60, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
        alto_fila = 20  # Alto de cada fila de la tabla
        tipo_col = ["t", "t", "t","t","n"]  # tipo de cada columna de la tabla
        alin_col = ["l", "l", "l","l","r"]  # tipo de cada columna de la tabla
        cantcar_col = [10, 12, 20,20,16]  # tipo de cada columna de la tabla
        dec_col = [0, 0, 0,0,3]  # tipo de cada columna de la tabla
       # Definir estilos de la tabla
        estilo_encabezado = ("Helvetica-Bold", 10, "white")  # Estilo de fuente para la fila de encabezado
        estilo_fila = ("Helvetica", 10)  # Estilo de fuente para las filas de datos
        color_encabezado = colors.gray  # Color de fondo para la fila de encabezado
        color_filas = [colors.lightgrey, colors.white]  # Colores de fondo para las filas de datos alternas
        color_celeste=(0.529, 0.808, 0.922)  # celeste en formato RGB
        e=0
        # Dibujar la tabla
        for i, fila in enumerate(data):
            if e > detmax:
                c.showPage()
                nropag = nropag+1
                e=0
                presupdeventacab_infpdftit(c, 680, nropag)
                fila_y=620

            # Calcular coprespadas de la fila actual
            fila_y = fila_y -  alto_fila
            # Dibujar las celdas de la fila
            fila_x = 20
            for j, valor in enumerate(fila):
                xlist = [fila_x-1, fila_x + ancho_col[j]]
                ylist = [fila_y, fila_y + alto_fila]
                c.grid(xlist, ylist)
                c.setFillColor(color_filas[i % 2])  # Establecer el color de fondo de la celda
                c.rect(fila_x, fila_y, ancho_col[j], alto_fila, fill=True, stroke=False)  # Dibujar la celda
                c.setFillColor(color_encabezado if i == 0 else colors.black)  # Establecer el color de fuente
                c.setFont("Helvetica", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades
                if tipo_col[j]=="n" :
                    total=total+valor
                    valor=str(formatnumver((valor),dec_col[j]))
                    c.setFont("Courier", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades

                valorw=formatearcampo(str(valor), int(cantcar_col[j]),alin_col[j], tipo_col[j])
                c.drawString(fila_x + 3, fila_y + alto_fila/4, str(valorw))  # Agregar el valor a la celda
                fila_x = fila_x + ancho_col[j]
                ultancho=ancho_col[j]
            e=e+1

        fila_y = fila_y - alto_fila
        valor = str(formatnumver((total), dec_col[j]))
        c.setFont("Courier", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades
        valorw = formatearcampo(str(valor), int(cantcar_col[j]), alin_col[j], tipo_col[j])

        c.drawString(fila_x -ultancho, fila_y + alto_fila / 4, str(valorw))  # Agregar el valor a la celda
        c.drawString(fila_x -ultancho-ultancho, fila_y + alto_fila / 4, str("TOTAL :"))  # Agregar el valor a la celda

        xlist = [fila_x-1-ultancho, fila_x]
        ylist = [fila_y, fila_y + alto_fila]
        c.grid(xlist, ylist)

def presupdeventadetinf_listar(request):
    idcliente = request.POST['idcliente']
    fechaini = request.POST['fechaini']
    fechafin = request.POST['fechafin']
    tipodoc = request.POST['tipodoc']
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab   = Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if idcliente == '':
        idcliente = 0
    if presupdeventacab .exists() and int(idcliente) >= 1:
        presupdeventacab   = presupdeventacab .filter(idcliente_id=idcliente)
    if presupdeventacab .exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab   = presupdeventacab .filter(fecha__gte=fecha_inicio)
    if presupdeventacab .exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab   = presupdeventacab .filter(fecha__lte=fecha_fin)
    if presupdeventacab .exists() and tipodoc:
        presupdeventacab   = presupdeventacab .filter(tipodoc=tipodoc)

    datos = []
    for objetocab in presupdeventacab:
        didpresupuestoventacab =objetocab.idpresupuestoventacab
        presupdeventadet = Presupuestoventadet.objects.filter(idpresupuestoventacab =didpresupuestoventacab).order_by('orden')
        for objetodet in presupdeventadet:
             datos.append({
            'fecha': objetocab.fecha.strftime("%d/%m/%Y"),
            'nropresupuesto': objetocab.nropresupuesto,
            'cliente': objetocab.cliente,
            'orden': objetodet.orden,
            'codigo': objetodet.codigo,
            'descripcion': objetodet.descripcion,
            'cantidad': objetodet.cantidad,
            'precio': formatnumver(objetodet.precio,0),
            'iva': formatnumver(objetodet.iva,0),
            'subtotal':formatnumver((objetodet.precio*objetodet.cantidad),0),
            'total':formatnumver(objetocab.total, 0)
            })


    total = sum([objeto.total * objeto.total for objeto in presupdeventacab ])
    Response = JsonResponse({'datos': datos})
    return Response


def presupdeventadetinfpdf(request):
    # -----------------------------------------------
    idcliente = request.GET.get('idcliente', '')
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    tipodoc = request.GET.get('tipodoc', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    presupdeventacab = Presupuestoventacab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if idcliente == '':
        idcliente = 0
    if presupdeventacab.exists() and int(idcliente) >= 1:
        presupdeventacab = presupdeventacab.filter(idcliente_id=idcliente)
    if presupdeventacab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        presupdeventacab = presupdeventacab.filter(fecha__gte=fecha_inicio)
    if presupdeventacab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        presupdeventacab = presupdeventacab.filter(fecha__lte=fecha_fin)
    if presupdeventacab.exists() and tipodoc:
        presupdeventacab = presupdeventacab.filter(tipodoc=tipodoc)

    # -------------------------------------------------

    #letter=8.5 * 11 pulgadas  215.9 *279.4 mm
    buf = io.BytesIO()
    #c = canvas.Canvas(buf)
    custom_oficio_width = 216 *2.7 # Anchura en unidades de medida (por ejemplo, puntos)
    custom_oficio_height = 330*2.1 # Altura en unidades de medida (por ejemplo, puntos)
    pagesize_oficio = portrait((custom_oficio_width, custom_oficio_height))
    c = canvas.Canvas(buf, pagesize=pagesize_oficio)
    #c = canvas.Canvas(buf, pagesize=legal)

    presupdeventadetinfpdftit(c,680,0)
    presupdeventadetinfpdfdet(presupdeventacab ,c,620)
    c.save()
    buf.seek(0)

    return FileResponse(buf, filename='venuepdf.pdf')

gancho_col = [80, 80, 100, 60, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
galto_fila = 20  # Alto de cada fila de la tabla
gtipo_col = ["t", "t", "t","t","n"]  # tipo de cada columna de la tabla
galin_col = ["l", "l", "l","l","r"]  # tipo de cada columna de la tabla
gcantcar_col = [10, 12, 20,20,16]  # tipo de cada columna de la tabla
gdec_col = [0, 0, 0,0,3]  # tipo de cada columna de la tabla

def presupdeventadetinfpdftit(c,y,npag):
        ahora = time.strftime("%c")
        fechahora = time.strftime("%c")
        fecha = time.strftime("%x")
        print(npag)

        #x = datetime.datetime.now()
        lx = 100
        hi = y
        c.setFont('Helvetica', 12)
        #c.drawString(lx+300, hi , "usuario:  fecha : " + fecha + "   " + str(x.hour) + ":" + str(
        #    x.minute) + "   pagina : " + str(nropag))
        c.setFont('Helvetica', 16)

         # Definir el color de relleno y el color de letra deseado
        color_relleno = (0.9, 0.9, 0.9)  # Utiliza los valores RGB del color de relleno que desees
        color_letra = (1, 0, 0)  # Utiliza los valores RGB del color de letra rojo
        color_relleno=(0, 0, 1)  # Azul en formato RGB
        color_letra=(1, 1, 1)  # blanco en formato RGB
        color_celeste=(0.529, 0.808, 0.922)  # celeste en formato RGB


        # Dibujar la cuadrícula
        ancho =150
        alto=30
        xlist = [lx, lx + ancho]
        ylist = [hi , hi + alto]
        # c.grid(xlist, ylist)

        # Establecer el color de relleno y dibujar el rectángulo en una celda específica
        c.setFillColor(colors.white)
        c.rect(lx+1, hi+1,  ancho-2, alto-2, fill=True, stroke=False)

        # Establecer el color de letra y dibujar el título en la celda
        #c.setFillColorRGB(*color_letra)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 22)  # Puedes ajustar la fuente y el tamaño según tus necesidades
        c.drawString(lx -10, hi-20,"INF. PRESUPUESTO  DE VENTA DETALLE ")
        presupdeventadetinfcabdet(c,620)


def presupdeventadetinfcabdet(c,y):
    data = [["ORD","CODIGO","DESCRIPCION","CANT","PRECIO","IVA","SUBTOTAL"]]

    # Definir coprespadas y dimensiones de la tabla
    fila_x = 20  # Coprespada x de la esquina superior izquierda de la tabla
    fila_y = y  # Coprespada y de la esquina superior izquierda de la tabla
    alto_fila = 20  # Alto de cada fila de la tabla
    ancho_col = [30, 50, 80, 40, 60, 30, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
    alto_fila = 20  # Alto de cada fila de la tabla
    tipo_col = [ "t", "t", "t", "t", "t","t", "t"]  # tipo de cada columna de la tabla
    alin_col = ["l", "l", "l", "l", "l", "l", "l"]  # tipo de cada columna de la tabla
    cantcar_col = [10, 12, 20, 20, 16, 20, 20]  # tipo de cada columna de la tabla

    c.setStrokeColor(colors.lightgrey)

    color_celeste = (0.529, 0.808, 0.922)  # celeste en formato RGB
    # Dibujar la tabla
    for i, fila in enumerate(data):
        for j, valor in enumerate(fila):
            xlist = [fila_x - 1, fila_x + ancho_col[j]]
            ylist = [fila_y, fila_y + alto_fila]
            c.grid(xlist, ylist)
            c.setFillColor(color_celeste)  # Establecer el color de fondo de la celda
            c.setFont("Helvetica", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades

            valorw = formatearcampo(str(valor), int(cantcar_col[j]), alin_col[j], tipo_col[j])
            c.drawString(fila_x + 3, fila_y + alto_fila / 4, str(valorw))  # Agregar el valor a la celda
            fila_x = fila_x + ancho_col[j]


def presupdeventadetinfpdfdet(presupdeventacab ,c,y):
        total=0
        detmax=25
        nropag=1
        # Definir coprespadas y dimensiones de la tabla
        fila_x = 30  # Coprespada x de la esquina superior izquierda de la tabla
        fila_y = y  # Coprespada y de la esquina superior izquierda de la tabla
        ancho_col = [30, 50, 80, 40, 60, 30, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
        alto_fila = 20  # Alto de cada fila de la tabla
        tipo_col = ["t", "t", "t", "n", "n", "n", "n"]  # tipo de cada columna de la tabla
        alin_col = [ "l", "l", "l", "r", "r", "r", "r"]  # tipo de cada columna de la tabla
        cantcar_col = [ 5, 20, 25, 7, 15, 5, 20]  # tipo de cada columna de la tabla
        dec_col = [ 0, 0, 0, 3, 0, 0, 0]  # tipo de cada columna de la tabla
        # Definir estilos de la tabla
        estilo_encabezado = ("Helvetica-Bold", 10, "white")  # Estilo de fuente para la fila de encabezado
        estilo_fila = ("Helvetica", 10)  # Estilo de fuente para las filas de datos
        color_encabezado = colors.gray  # Color de fondo para la fila de encabezado
        color_filas = [colors.lightgrey, colors.white]  # Colores de fondo para las filas de datos alternas
        color_celeste = (0.529, 0.808, 0.922)  # celeste en formato RGB
        e=0

        for objetocab in presupdeventacab :
            didpresupuestoventacab = objetocab.idpresupuestoventacab
            presupdeventadet = Presupuestoventadet.objects.filter(idpresupuestoventacab =didpresupuestoventacab).order_by('orden')
            if presupdeventadet.exists():
                fila_x=30
                fila_y = fila_y - alto_fila
                celda(c,"FECHA :", fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                fila_x=fila_x+50
                celda(c, objetocab.fecha.strftime('%d/%m/%Y'), fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                fila_x=fila_x+50
                celda(c,"FACTURA :", fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                fila_x=fila_x+50
                celda(c, objetocab.nropresupuesto, fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                fila_x=fila_x+60
                celda(c,"CLIENTE :", fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                fila_x=fila_x+60
                celda(c, objetocab.cliente, fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                prespmax=0
                data = []
                for objetodet in presupdeventadet:
                    prespmax=objetodet.orden
                    nueva_fila = [
                      str(objetodet.orden),
                      str(objetodet.codigo),
                      str(objetodet.descripcion),
                      objetodet.cantidad,
                      objetodet.precio,
                      objetodet.iva,
                      (objetodet.precio*objetodet.cantidad)
                    ]
                    data.append(nueva_fila)
                e=e+2
                # Dibujar la tabla
                i=0
                total=0
                for i, fila in enumerate(data):
                    e = e + 1
                    if e > detmax:
                        c.showPage()
                        nropag = nropag+1
                        e=0
                        presupdeventadetinfpdftit(c, 680, nropag)
                        fila_y=620

                    # Calcular coprespadas de la fila actual
                    fila_y = fila_y -  alto_fila
                    # Dibujar las celdas de la fila
                    fila_x = 20
                    j=0
                    for j, valor in enumerate(fila):
                        xlist = [fila_x-1, fila_x + ancho_col[j]]
                        ylist = [fila_y, fila_y + alto_fila]
                        # Set the color of the grid lines
                        c.setStrokeColor(colors.lightgrey)
                        # Set the color of the margin area
                        c.grid(xlist, ylist)
                        c.setFillColor(color_filas[i % 2])  # Establecer el color de fondo de la celda
                        #c.rect(fila_x, fila_y, ancho_col[j], alto_fila, fill=True, stroke=False)  # Dibujar la celda
                        c.setFillColor(colors.black)  # Establecer el color de fuente
                        c.setFont("Helvetica", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades
                        if tipo_col[j]=="n":
                            if j == 6:
                                total = total + valor
                            valor = str(formatnumver((valor), dec_col[j]))
                            c.setFont("Courier", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades

                        valorw=formatearcampo(str(valor), int(cantcar_col[j]),alin_col[j], tipo_col[j])
                        c.drawString(fila_x + 3, fila_y + alto_fila/4, str(valorw))  # Agregar el valor a la celda
                        fila_x = fila_x + ancho_col[j]
                        if prespmax-1 == i and j==6:
                            i=i+1

                            fila_x=fila_x-ancho_col[j]
                            fila_y = fila_y - alto_fila

                            xlist = [fila_x - 1, fila_x + ancho_col[j]]
                            ylist = [fila_y, fila_y + alto_fila]
                            c.grid(xlist, ylist)

                            valor = str(formatnumver((total), 0))
                            c.setFont("Courier", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades
                            #xlist = [fila_x - 1, fila_x + ancho_col[j]]
                            valorw = formatearcampo(str(valor), int(cantcar_col[j]), alin_col[j], tipo_col[j])
                            c.drawString(fila_x + 3, fila_y + alto_fila / 4, str(valorw))  # Agregar el valor a la celda


def celda(c,valor,fila_x,fila_y,iancho_col,ialto_fila,itipo_col,ialin_col,icantcar_col,idec_col):

    xlist = [fila_x - 1, fila_x + iancho_col]
    ylist = [fila_y, fila_y + ialto_fila]
    #c.grid(xlist, ylist)
    c.setFillColor(colors.black)  # Establecer el color de fuente

    #c.rect(fila_x, fila_y, iancho_col, ialto_fila, fill=True, stroke=False)  # Dibujar la celda
    c.setFont("Helvetica", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades
    if itipo_col == "n":
        valor = str(formatnumver((valor), idec_col))
        c.setFont("Courier", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades
    valorw = formatearcampo(str(valor), icantcar_col, ialin_col, itipo_col)
    c.drawString(fila_x + 3, fila_y + ialto_fila / 4, str(valorw))  # Agregar el valor a la celda
