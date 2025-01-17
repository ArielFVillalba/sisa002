from django.db import connections
from django.shortcuts import render, HttpResponse, redirect , get_object_or_404, redirect
from django.urls import reverse_lazy
from django.utils.datetime_safe import datetime
from django.views.generic import CreateView, UpdateView, DeleteView
from openpyxl.compat import numbers
from openpyxl.styles import Alignment

from .forms import *
from django.contrib import messages
from reportlab.lib import colors
from django.http import JsonResponse, FileResponse, HttpResponse
from django.shortcuts import render
from inicio.funcion import *
from django.shortcuts import redirect
import io
import csv
from reportlab.pdfgen import canvas
import time
import datetime
from datetime import datetime

from reportlab.lib.pagesizes import letter, landscape, inch, A4, A3, legal, B4, B5, A5, portrait
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from django.http import HttpResponse


def movdepcabinf_listar(request):
    if not request.user.is_authenticated:
        return redirect('login')

    fechaini = request.POST['fechaini']
    fechafin = request.POST['fechafin']
    depsalida = request.POST['depsalida']
    depentrada = request.POST['depentrada']
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


       
    movdepcab = Movdepcab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depentrada__contains=depentrada,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depsalida__contains=depsalida,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)

    datos = []
    for objeto in movdepcab:
        datos.append({
        'fecha': objeto.fecha.strftime("%d/%m/%Y"),
        'nromov': objeto.nromov,
        'depsalida': objeto.depsalida,
        'depentrada': objeto.depentrada,
    })

    #total = sum([objeto.total * objeto.total for objeto in compracab])

    Response = JsonResponse({'datos': datos})
    return Response


def movdepcabinf(request):
    if not request.user.is_authenticated:
        return redirect('login')
    return render(request, 'stock_inf/lstmovdeposito.html')



def movdepdetinfcsv(request):
    if not request.user.is_authenticated:
        return redirect('login')
    # -----------------------------------------------
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    depsalida = request.GET.get('depsalida', '')
    depentrada = request.GET.get('depentrada', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


       

    movdepcab = Movdepcab.objects.all()
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depentrada__contains=depentrada,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depsalida__contains=depsalida,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)

    # -------------------------------------------------



    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="datos.csv"'
    writer = csv.writer(response, delimiter=';')  # Establece el separador como ";"
    writer.writerow(["FECHA", "NRO MOV","ORDEN","codigo","descripcion" ,"cantidad","DEP SALIDA","DEP ENTRADA","USUARIO"])
    total=0
    for objetocab in movdepcab:
        idmovdepcab = objetocab.idmovdepcab
        compradet = Movdepdet.objects.filter(idmovdepcab=idmovdepcab).order_by('orden')

        for objetodet in compradet:
            #if objetodet.orden==1 :
                  #  writer.writerow(["","","","","","","","","TOTAL",formatnumver(total,0)])
                  #  total = 0
                  #  writer.writerow([""])
            #total=total+(objetodet.cantidad * objetodet.precio)

            writer.writerow([
                objetocab.fecha.strftime("%d/%m/%Y"),
                objetocab.nromov,
                objetodet.orden,
                objetodet.codigo,
                objetodet.descripcion,
                formatnumver(objetodet.cantidad,3),
                objetodet.depsalida,
                objetodet.depentrada,
                objetodet.usuario,

            ])
    #writer.writerow(["", "", "", "", "", "", "", "", "TOTAL", formatnumver(total, 0)])

    return response

def movdepdetinfcsvee(request):
    if not request.user.is_authenticated:
        return redirect('login')
    # -----------------------------------------------
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    depsalida = request.GET.get('depsalida', '')
    depentrada = request.GET.get('depentrada', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


       

    movdepcab = Movdepcab.objects.all(idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depentrada__contains=depentrada,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depsalida__contains=depsalida,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)

    # -------------------------------------------------

    objetos=movdepcab

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="datos.csv"'
    writer = csv.writer(response, delimiter=';')  # Establece el separador como ";"
    writer.writerow(["FECHA", "NRO MOV", "DEP SALIDA", "DEP ENTRADA","TOTAL"])
    # Agrega los datos de cada fila
    for fila in objetos:
        writer.writerow([fila.fecha, fila.nromov,fila.depsalida, fila.depentrada])
    return response


def movdepcabinfcsv(request):
    if not request.user.is_authenticated:
        return redirect('login')
    # -----------------------------------------------
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    depsalida = request.GET.get('depsalida', '')
    depentrada = request.GET.get('depentrada', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


       

    movdepcab = Movdepcab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depentrada__contains=depentrada,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depsalida__contains=depsalida,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)

    # -------------------------------------------------


    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="datos.csv"'
    writer = csv.writer(response, delimiter=';')  # Establece el separador como ";"
    writer.writerow(["FECHA", "Nro Mov", "ORDEN","codigo","descripcion" ,"cantidad","DEP. SALIDA","DEP ENTRADA"])
    total=0
    for objetocab in movdepcab:
        idmovdepcab = objetocab.idmovdepcab
        movdepdet = Movdepdet.objects.filter(idmovdepcab=idmovdepcab).order_by('orden')

        for objetodet in movdepdet:
            #if objetodet.orden==1 :
                   # writer.writerow(["","","","","","","","","TOTAL",formatnumver(total,0)])
                    #total = 0
                    #writer.writerow([""])
            #total=total+(objetodet.cantidad * objetodet.precio)

            writer.writerow([
                objetocab.fecha.strftime("%d/%m/%Y"),
                objetocab.nromov,
                objetodet.orden,
                objetodet.codigo,
                objetodet.descripcion,
                formatnumver(objetodet.cantidad,3),
                objetodet.depsalida,
                objetodet.depentrada,
                objetodet.usuario,

            ])
   # writer.writerow(["", "", "", "", "", "", "", "", "TOTAL", formatnumver(total, 0)])

    return response



def movdepcabinfexcel(request):
    if not request.user.is_authenticated:
        return redirect('login')
    # -----------------------------------------------
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    depsalida = request.GET.get('depsalida', '')
    depentrada = request.GET.get('depentrada', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


    movdepcab = Movdepcab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depentrada__contains=depentrada,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depsalida__contains=depsalida,idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)

    # -------------------------------------------------
    data = movdepcab

    # Crear un nuevo libro de trabajo (workbook)
    wb = Workbook()

    # Obtén tus datos de Django y genera las filas del archivo Excel
    # Ejemplo ficticio con datos de un modelo llamado "MiModelo"

    # Crea una hoja de cálculo en el libro de trabajo
    sheet = wb.active
    sheet.title = "LISTADO MOV DEPOSITO"

    # Agrega encabezados a la hoja de cálculo
    # headers = ['Campo1', 'Campo2', 'Campo3']
    headers=["FECHA", "NRO MOVIMIENTO", "DEP SALIDA", "DEP ENTRADA"]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}2"]
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alinea el título al centro
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15


    # Fusiona las celdas para el título general
    inicio_celda = sheet.cell(row=1, column=1)
    fin_celda = sheet.cell(row=1, column=len(headers))
    sheet.merge_cells(start_row=1, start_column=inicio_celda.column, end_row=1, end_column=fin_celda.column)
    inicio_celda.alignment = Alignment(horizontal='center')
    # Establece el valor del título general en la celda fusionada
    titulo_general = "LISTADO MOV DEPOSITO"
    inicio_celda.value = titulo_general

    # Agrega los datos de cada fila a la hoja de cálculo
    for row_num, fila in enumerate(data, 3):
        sheet.cell(row=row_num, column=1, value=fila.fecha.strftime('%d/%m/%Y'))
        sheet.cell(row=row_num, column=2, value=fila.nromov)
        sheet.cell(row=row_num, column=3, value=fila.depsalida)
        sheet.cell(row=row_num, column=4, value=fila.depentrada)
        #total_cell = sheet.cell(row=row_num, column=5, value=fila.total)


    # Configura el tipo de respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="datos.xlsx"'

    # Guarda el libro de trabajo en la respuesta HTTP
    wb.save(response)

    return response

def movdepdetinfexcel(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # -----------------------------------------------
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    depentrada = request.GET.get('depentrada', '')
    depsalida = request.GET.get('depsalida', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


       

    movdepcab = Movdepcab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depentrada__contains=depentrada,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() :
        movdepcab = movdepcab.filter(depsalida__contains=depsalida,idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)

    # -------------------------------------------------
    data = movdepcab

    # Crear un nuevo libro de trabajo (workbook)
    wb = Workbook()

    # Obtén tus datos de Django y genera las filas del archivo Excel
    # Ejemplo ficticio con datos de un modelo llamado "MiModelo"

    # Crea una hoja de cálculo en el libro de trabajo
    sheet = wb.active
    sheet.title = "Listado MOV DEPOSITO"

    # Agrega encabezados a la hoja de cálculo
    # headers = ['Campo1', 'Campo2', 'Campo3']
    headers=["FECHA", "NRO MOV", "ORDEN","CODIGO", "DESCRIPCION", "CANTIDAD","DEP SALIDA", "DEP ENTRADA","USUARIO"]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}2"]
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Alinea el título al centro
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 15
        sheet.column_dimensions['G'].width = 15
        sheet.column_dimensions['H'].width = 15
        sheet.column_dimensions['I'].width = 15


    # Fusiona las celdas para el título general
    inicio_celda = sheet.cell(row=1, column=1)
    fin_celda = sheet.cell(row=1, column=len(headers))
    sheet.merge_cells(start_row=1, start_column=inicio_celda.column, end_row=1, end_column=fin_celda.column)
    inicio_celda.alignment = Alignment(horizontal='center')
    # Establece el valor del título general en la celda fusionada
    titulo_general = "LISTADO MOV DEPOSITO DETALLADO  "
    inicio_celda.value = titulo_general
    row_num=3
    total=0
    for objetocab in movdepcab:
        idmovdepcab = objetocab.idmovdepcab
        movdepdet = Movdepdet.objects.filter(idmovdepcab=idmovdepcab).order_by('orden')
        if total>0:
            row_num=row_num+1
            sheet.cell(row=row_num, column=10, value=formatnumver(total,0))
            total=0
            row_num=row_num+1



        for objetodet in movdepdet:
            #total=total+(objetodet.cantidad * objetodet.precio)
            row_num=row_num+1
            sheet.cell(row=row_num, column=1, value=objetocab.fecha.strftime('%d/%m/%Y'))
            sheet.cell(row=row_num, column=2, value=objetocab.nromov)
            sheet.cell(row=row_num, column=3, value=objetodet.orden)
            sheet.cell(row=row_num, column=4, value=objetodet.codigo)
            sheet.cell(row=row_num, column=5, value=objetodet.descripcion)
            sheet.cell(row=row_num, column=6, value=objetodet.cantidad)
            sheet.cell(row=row_num, column=7, value=objetodet.depsalida)
            sheet.cell(row=row_num, column=8, value=objetodet.depentrada)
            sheet.cell(row=row_num, column=9, value=objetodet.usuario)

    row_num = row_num + 1
   # sheet.cell(row=row_num, column=10, value=formatnumver(total, 0))

    # Configura el tipo de respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="datos.xlsx"'

    # Guarda el libro de trabajo en la respuesta HTTP
    wb.save(response)

    return response

def movdepcabinfpdf(request):
    if not request.user.is_authenticated:
        return redirect('login')

    # -----------------------------------------------
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    depentrada = request.GET.get('depentrada', '')
    depsalida = request.GET.get('depsalida', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


       

    movdepcab = Movdepcab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists():
        movdepcab = movdepcab.filter(depentrada__contains=depentrada,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists():
        movdepcab = movdepcab.filter(depsalida__contains=depsalida,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)


    # -------------------------------------------------

    #letter=8.5 * 11 pulgadas  215.9 *279.4 mm
    buf = io.BytesIO()
    #c = canvas.Canvas(buf)
    custom_oficio_width = 216 *2.7 # Anchura en unidades de medida (por ejemplo, puntos)
    custom_oficio_height = 330*2.1 # Altura en unidades de medida (por ejemplo, puntos)
    pagesize_oficio = portrait((custom_oficio_width, custom_oficio_height))
    c = canvas.Canvas(buf, pagesize=pagesize_oficio)
    #c = canvas.Canvas(buf, pagesize=legal)

    movdepcabinfpdftit(c,680,0)
    movdepcabinfpdfdet(movdepcab,c,620)
    c.save()
    buf.seek(0)

    return FileResponse(buf, filename='venuepdf.pdf')

gancho_col = [80, 80, 100, 60, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
galto_fila = 20  # Alto de cada fila de la tabla
gtipo_col = ["t", "t", "t","t","n"]  # tipo de cada columna de la tabla
galin_col = ["l", "l", "l","l","r"]  # tipo de cada columna de la tabla
gcantcar_col = [10, 12, 20,20,16]  # tipo de cada columna de la tabla
gdec_col = [0, 0, 0,0,3]  # tipo de cada columna de la tabla

def movdepcabinfpdftit(c,y,npag):
        ahora = time.strftime("%c")
        fechahora = time.strftime("%c")
        fecha = time.strftime("%x")
        print(npag)

        #x = datetime.datetime.now()
        lx = 100
        hi = y
        c.setFont('Helvetica', 12)
        #c.drawString(lx+300, hi , "usuario:  fecha : " + fecha + "   " + str(x.hour) + ":" + str(
        #    x.minute) + "   pagina : " + str(nropag))
        c.setFont('Helvetica', 16)

         # Definir el color de relleno y el color de letra deseado
        color_relleno = (0.9, 0.9, 0.9)  # Utiliza los valores RGB del color de relleno que desees
        color_letra = (1, 0, 0)  # Utiliza los valores RGB del color de letra rojo
        color_relleno=(0, 0, 1)  # Azul en formato RGB
        color_letra=(1, 1, 1)  # blanco en formato RGB
        color_celeste=(0.529, 0.808, 0.922)  # celeste en formato RGB


        # Dibujar la cuadrícula
        ancho =150
        alto=30
        xlist = [lx, lx + ancho]
        ylist = [hi , hi + alto]
        # c.grid(xlist, ylist)

        # Establecer el color de relleno y dibujar el rectángulo en una celda específica
        c.setFillColor(colors.white)
        c.rect(lx+1, hi+1,  ancho-2, alto-2, fill=True, stroke=False)

        # Establecer el color de letra y dibujar el título en la celda
        #c.setFillColorRGB(*color_letra)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 22)  # Puedes ajustar la fuente y el tamaño según tus necesidades
        c.drawString(lx + 20, hi-20,"INFORME MOVIMIENTO DEPOSITO ")
        movdepcabinfpdfcabdet(c,620)


def movdepcabinfpdfcabdet(c,y):
    data = [["FECHA", "NRO MOV", "DEP SALIDA", "DEP ENTRADA"]]

    # Definir coordenadas y dimensiones de la tabla
    fila_x = 20  # Coordenada x de la esquina superior izquierda de la tabla
    fila_y = y  # Coordenada y de la esquina superior izquierda de la tabla
    alto_fila = 20  # Alto de cada fila de la tabla
    ancho_col = [100, 100, 100, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
    tipo_col = ["t", "t", "t", "t", "t"]  # tipo de cada columna de la tabla
    alin_col = ["l", "l", "l", "l", "l"]  # tipo de cada columna de la tabla
    cantcar_col = [10, 20, 20, 20 ] # tipo de cada columna de la tabla
    color_celeste = (0.529, 0.808, 0.922)  # celeste en formato RGB
    # Dibujar la tabla
    for i, fila in enumerate(data):
        for j, valor in enumerate(fila):
            xlist = [fila_x - 1, fila_x + ancho_col[j]]
            ylist = [fila_y, fila_y + alto_fila]
            c.grid(xlist, ylist)
            c.setFillColor(color_celeste)  # Establecer el color de fondo de la celda
            c.setFont("Helvetica", 12)  # Puedes ajustar la fuente y el tamaño según tus necesidades

            valorw = formatearcampo(str(valor), int(cantcar_col[j]), alin_col[j], tipo_col[j])
            c.drawString(fila_x + 3, fila_y + alto_fila / 4, str(valorw))  # Agregar el valor a la celda
            fila_x = fila_x + ancho_col[j]


def movdepcabinfpdfdet(compracab,c,y):
        total=0
        detmax=25
        nropag=1
        #data = [["FECHA", "NRO MOV", "DEP SALIDA", "DEP ENTRADA"]]
        data = []
        objetos = compracab
        for objeto in objetos:
                nueva_fila = [str(objeto.fecha.strftime('%d/%m/%Y')),
                str(objeto.nromov),
                str(objeto.depsalida),
                str(objeto.depentrada),
                ]
                data.append(nueva_fila)

        # Definir coordenadas y dimensiones de la tabla
        fila_x = 30  # Coordenada x de la esquina superior izquierda de la tabla
        fila_y = y  # Coordenada y de la esquina superior izquierda de la tabla
        ancho_col = [100, 100, 100, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
        alto_fila = 20  # Alto de cada fila de la tabla
        tipo_col = ["t", "t", "t","t"]  # tipo de cada columna de la tabla
        alin_col = ["l", "l", "l","l","r"]  # tipo de cada columna de la tabla
        cantcar_col = [10, 20, 20,20]  # tipo de cada columna de la tabla
        dec_col = [0, 0, 0,0]  # tipo de cada columna de la tabla
       # Definir estilos de la tabla
        estilo_encabezado = ("Helvetica-Bold", 10, "white")  # Estilo de fuente para la fila de encabezado
        estilo_fila = ("Helvetica", 10)  # Estilo de fuente para las filas de datos
        color_encabezado = colors.gray  # Color de fondo para la fila de encabezado
        color_filas = [colors.lightgrey, colors.white]  # Colores de fondo para las filas de datos alternas
        color_celeste=(0.529, 0.808, 0.922)  # celeste en formato RGB
        e=0
        # Dibujar la tabla
        for i, fila in enumerate(data):
            if e > detmax:
                c.showPage()
                nropag = nropag+1
                e=0
                movdepcabinfpdftit(c, 680, nropag)
                fila_y=620

            # Calcular coordenadas de la fila actual
            fila_y = fila_y -  alto_fila
            # Dibujar las celdas de la fila
            fila_x = 20
            for j, valor in enumerate(fila):
                xlist = [fila_x-1, fila_x + ancho_col[j]]
                ylist = [fila_y, fila_y + alto_fila]
                c.grid(xlist, ylist)
                c.setFillColor(color_filas[i % 2])  # Establecer el color de fondo de la celda
                c.rect(fila_x, fila_y, ancho_col[j], alto_fila, fill=True, stroke=False)  # Dibujar la celda
                c.setFillColor(color_encabezado if i == 0 else colors.black)  # Establecer el color de fuente
                c.setFont("Helvetica", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades
                if tipo_col[j]=="n" :
                    total=total+valor
                    valor=str(formatnumver((valor),dec_col[j]))
                    c.setFont("Courier", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades

                valorw=formatearcampo(str(valor), int(cantcar_col[j]),alin_col[j], tipo_col[j])
                c.drawString(fila_x + 3, fila_y + alto_fila/4, str(valorw))  # Agregar el valor a la celda
                fila_x = fila_x + ancho_col[j]
                ultancho=ancho_col[j]
            e=e+1

        #fila_y = fila_y - alto_fila
        #valor = str(formatnumver((total), dec_col[j]))
        #c.setFont("Courier", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades
        #valorw = formatearcampo(str(valor), int(cantcar_col[j]), alin_col[j], tipo_col[j])

        #c.drawString(fila_x -ultancho, fila_y + alto_fila / 4, str(valorw))  # Agregar el valor a la celda
        #c.drawString(fila_x -ultancho-ultancho, fila_y + alto_fila / 4, str("TOTAL :"))  # Agregar el valor a la celda

        #xlist = [fila_x-1-ultancho, fila_x]
        #ylist = [fila_y, fila_y + alto_fila]
        #c.grid(xlist, ylist)

def movdepdetinf_listar(request):

    fechaini = request.POST['fechaini']
    fechafin = request.POST['fechafin']
    depsalida = request.POST['depsalida']
    depentrada = request.POST['depentrada']
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')

    movdepcab = Movdepcab.objects.all(idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and depsalida:
        compracab = movdepcab.filter(depsalida=depsalida,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and depsalida:
        depentrada = movdepcab.filter(depentrada=depentrada,idempresa=didempresa,idsucursal=didsucursal)


    datos = []
    for objetocab in movdepcab:
        idmovdepcab =objetocab.idmovdepcab
        movdepdet = Movdepdet.objects.filter(idmovdepcab=idmovdepcab).order_by('orden')
        for objetodet in movdepdet:
             datos.append({
            'fecha': objetocab.fecha.strftime("%d/%m/%Y"),
            'nromov': objetocab.nromov,
            'orden': objetodet.orden,
            'codigo': objetodet.codigo,
            'descripcion': objetodet.descripcion,
            'cantidad': objetodet.cantidad,
            'depsalida': objetodet.depsalida,
            'depentrada': objetodet.depentrada,
            'usuario':objetodet.usuario,
            })


    #total = sum([objeto.total * objeto.total for objeto in compracab])
    Response = JsonResponse({'datos': datos})
    return Response


def movdepdetinfpdf(request):
    # -----------------------------------------------
    fechaini = request.GET.get('fechaini', '')
    fechafin = request.GET.get('fechafin', '')
    depsalida = request.GET.get('depsalida', '')
    depentrada = request.GET.get('depentrada', '')
    idempresa = request.session.get('idempresa')
    idsucursal = request.session.get('idsucursal')
    didempresa = desencriptar_datos2(idempresa, request)
    didsucursal = desencriptar_datos2(idsucursal, request)
    if int(didempresa) < 1 or int(didsucursal) < 1:
        return redirect('login')


       

    movdepcab = Movdepcab.objects.filter(idempresa=didempresa,idsucursal=didsucursal)

    if movdepcab.exists() and fechaini:
        fecha_inicio = datetime.strptime(fechaini, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__gte=fecha_inicio,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and fechafin:
        fecha_fin = datetime.strptime(fechafin, '%Y-%m-%d').date()
        movdepcab = movdepcab.filter(fecha__lte=fecha_fin,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and depsalida:
        movdepcab = movdepcab.filter(depsalida=depsalida,idempresa=didempresa,idsucursal=didsucursal)
    if movdepcab.exists() and depsalida:
        movdepcab = movdepcab.filter(depentrada=depentrada,idempresa=didempresa,idsucursal=didsucursal)

    # -------------------------------------------------

    #letter=8.5 * 11 pulgadas  215.9 *279.4 mm
    buf = io.BytesIO()
    #c = canvas.Canvas(buf)
    custom_oficio_width = 216 *2.7 # Anchura en unidades de medida (por ejemplo, puntos)
    custom_oficio_height = 330*2.1 # Altura en unidades de medida (por ejemplo, puntos)
    pagesize_oficio = portrait((custom_oficio_width, custom_oficio_height))
    c = canvas.Canvas(buf, pagesize=pagesize_oficio)
    #c = canvas.Canvas(buf, pagesize=legal)

    movdepdetinfpdftit(c,680,0)
    movdepdetinfpdfdet(movdepcab,c,620)
    c.save()
    buf.seek(0)

    return FileResponse(buf, filename='venuepdf.pdf')

gancho_col = [60, 80, 100, 60, 100]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
galto_fila = 20  # Alto de cada fila de la tabla
gtipo_col = ["t", "t", "t","t","t","t","t","t"]  # tipo de cada columna de la tabla
galin_col = ["l", "l", "l","l","l","l","l","l"]  # tipo de cada columna de la tabla
gcantcar_col = [10, 20, 20,20,12,10,10]  # tipo de cada columna de la tabla
gdec_col = [0, 0, 0,0,0,0,0,0]  # tipo de cada columna de la tabla

def movdepdetinfpdftit(c,y,npag):
        ahora = time.strftime("%c")
        fechahora = time.strftime("%c")
        fecha = time.strftime("%x")
        print(npag)

        #x = datetime.datetime.now()
        lx = 100
        hi = y
        c.setFont('Helvetica', 12)
        #c.drawString(lx+300, hi , "usuario:  fecha : " + fecha + "   " + str(x.hour) + ":" + str(
        #    x.minute) + "   pagina : " + str(nropag))
        c.setFont('Helvetica', 16)

         # Definir el color de relleno y el color de letra deseado
        color_relleno = (0.9, 0.9, 0.9)  # Utiliza los valores RGB del color de relleno que desees
        color_letra = (1, 0, 0)  # Utiliza los valores RGB del color de letra rojo
        color_relleno=(0, 0, 1)  # Azul en formato RGB
        color_letra=(1, 1, 1)  # blanco en formato RGB
        color_celeste=(0.529, 0.808, 0.922)  # celeste en formato RGB


        # Dibujar la cuadrícula
        ancho =150
        alto=30
        xlist = [lx, lx + ancho]
        ylist = [hi , hi + alto]
        # c.grid(xlist, ylist)

        # Establecer el color de relleno y dibujar el rectángulo en una celda específica
        c.setFillColor(colors.white)
        c.rect(lx+1, hi+1,  ancho-2, alto-2, fill=True, stroke=False)

        # Establecer el color de letra y dibujar el título en la celda
        #c.setFillColorRGB(*color_letra)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 22)  # Puedes ajustar la fuente y el tamaño según tus necesidades
        c.drawString(lx + 20, hi-20,"INFORME DEP MOVIMIENTO DETALLE ")
        movdepdetinfcabdet(c,620)


def movdepdetinfcabdet(c,y):
    data = [["FECHA","NRO MOV.","ORD","CODIGO","DESCRIP","CANT","DEP. SAL","DEP ENT"]]

    # Definir coordenadas y dimensiones de la tabla
    fila_x = 20  # Coordenada x de la esquina superior izquierda de la tabla
    fila_y = y  # Coordenada y de la esquina superior izquierda de la tabla
    alto_fila = 20  # Alto de cada fila de la tabla
    ancho_col = [50, 80, 40, 100, 120, 50, 50, 50]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
    alto_fila = 20  # Alto de cada fila de la tabla
    tipo_col = [ "t", "t", "t", "t", "t","t", "t", "t"]  # tipo de cada columna de la tabla
    alin_col = ["l", "l", "l", "l", "l", "l", "l", "l"]  # tipo de cada columna de la tabla
    cantcar_col = [10, 20, 5, 20, 10, 10, 10, 10]  # tipo de cada columna de la tabla

    c.setStrokeColor(colors.lightgrey)

    color_celeste = (0.529, 0.808, 0.922)  # celeste en formato RGB
    # Dibujar la tabla
    for i, fila in enumerate(data):
        for j, valor in enumerate(fila):
            xlist = [fila_x - 1, fila_x + ancho_col[j]]
            ylist = [fila_y, fila_y + alto_fila]
            c.grid(xlist, ylist)
            c.setFillColor(color_celeste)  # Establecer el color de fondo de la celda
            c.setFont("Helvetica", 10)  # Puedes ajustar la fuente y el tamaño según tus necesidades

            valorw = formatearcampo(str(valor), int(cantcar_col[j]), alin_col[j], tipo_col[j])
            c.drawString(fila_x + 3, fila_y + alto_fila / 4, str(valorw))  # Agregar el valor a la celda
            fila_x = fila_x + ancho_col[j]


def movdepdetinfpdfdet(movdepcab,c,y):
        total=0
        detmax=25
        nropag=1
        # Definir coordenadas y dimensiones de la tabla
        fila_x = 30  # Coordenada x de la esquina superior izquierda de la tabla
        fila_y = y  # Coordenada y de la esquina superior izquierda de la tabla
        ancho_col = [50, 80, 40, 100, 120, 50, 50, 50]  # Ancho de cada columna de la tabla 100 3.5 cm 20 7 cm
        alto_fila = 20  # Alto de cada fila de la tabla
        tipo_col = ["t", "t", "t", "t", "t", "t", "t", "t"]  # tipo de cada columna de la tabla
        alin_col = ["l", "l", "l", "l", "l", "l", "l", "l"]  # tipo de cada columna de la tabla
        cantcar_col = [10, 20, 5, 20, 10, 10, 10, 10]  # tipo de cada columna de la tabla
        # Definir estilos de la tabla
        estilo_encabezado = ("Helvetica-Bold", 10, "white")  # Estilo de fuente para la fila de encabezado
        estilo_fila = ("Helvetica", 10)  # Estilo de fuente para las filas de datos
        color_encabezado = colors.gray  # Color de fondo para la fila de encabezado
        color_filas = [colors.lightgrey, colors.white]  # Colores de fondo para las filas de datos alternas
        color_celeste = (0.529, 0.808, 0.922)  # celeste en formato RGB
        e=0

        for objetocab in movdepcab:
            idmovdepcab = objetocab.idmovdepcab
            movdepdet = Movdepdet.objects.filter(idmovdepcab=idmovdepcab).order_by('orden')
            if movdepdet.exists():
                print(" movdepdet " + str( movdepdet ))

                fila_x=30
                fila_y = fila_y - alto_fila
                #celda(c,"FECHA :", fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                #fila_x=fila_x+50
                #celda(c, objetocab.fecha.strftime('%d/%m/%Y'), fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                fila_x=fila_x+50
                #celda(c,"NRO MOV :", fila_x, fila_y, 10, alto_fila,"t", "l", 20, 0)
                fila_x=fila_x+50
                ordenmax=0
                data = []
                for objetodet in movdepdet:
                    ordenmax=objetodet.orden
                    nueva_fila = [
                        str(objetocab.fecha),
                        str(objetocab.nromov),
                        str(objetodet.orden),
                        str(objetodet.codigo),
                        str(objetodet.descripcion),
                        objetodet.cantidad,
                        objetodet.depsalida,
                        objetodet.depentrada
                    ]
                    data.append(nueva_fila)
                e=e+2
                # Dibujar la tabla
                i=0
                total=0
                for i, fila in enumerate(data):
                    e = e + 1
                    if e > detmax:
                        c.showPage()
                        nropag = nropag+1
                        e=0
                        movdepdetinfpdftit(c, 680, nropag)
                        fila_y=620
                    # Calcular coordenadas de la fila actual
                    fila_y = fila_y -  alto_fila
                    # Dibujar las celdas de la fila
                    fila_x = 20
                    j=0
                    for j, valor in enumerate(fila):
                        print(" fila "+ str(fila))
                        print(" j "+ str(j))

                        xlist = [fila_x-1, fila_x + ancho_col[j]]
                        ylist = [fila_y, fila_y + alto_fila]
                        # Set the color of the grid lines
                        c.setStrokeColor(colors.lightgrey)
                        # Set the color of the margin area
                        c.grid(xlist, ylist)
                        c.setFillColor(color_filas[i % 2])  # Establecer el color de fondo de la celda
                        #c.rect(fila_x, fila_y, ancho_col[j], alto_fila, fill=True, stroke=False)  # Dibujar la celda
                        c.setFillColor(colors.black)  # Establecer el color de fuente
                        c.setFont("Helvetica", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades
                        valorw=formatearcampo(str(valor), int(cantcar_col[j]),alin_col[j], tipo_col[j])
                        c.drawString(fila_x + 3, fila_y + alto_fila/4, str(valorw))  # Agregar el valor a la celda
                        fila_x = fila_x + ancho_col[j]

def celda(c,valor,fila_x,fila_y,iancho_col,ialto_fila,itipo_col,ialin_col,icantcar_col,idec_col):

    xlist = [fila_x - 1, fila_x + iancho_col]
    ylist = [fila_y, fila_y + ialto_fila]
    #c.grid(xlist, ylist)
    c.setFillColor(colors.black)  # Establecer el color de fuente

    #c.rect(fila_x, fila_y, iancho_col, ialto_fila, fill=True, stroke=False)  # Dibujar la celda
    c.setFont("Helvetica", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades
    if itipo_col == "n":
        valor = str(formatnumver((valor), idec_col))
        c.setFont("Courier", 8)  # Puedes ajustar la fuente y el tamaño según tus necesidades
    valorw = formatearcampo(str(valor), icantcar_col, ialin_col, itipo_col)
    c.drawString(fila_x + 3, fila_y + ialto_fila / 4, str(valorw))  # Agregar el valor a la celda
